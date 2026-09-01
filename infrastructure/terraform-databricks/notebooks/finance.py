# Databricks notebook source
# MAGIC %pip install --quiet openpyxl

# COMMAND ----------

"""`finance` task of job ow_tp_custbill: conversion of etl/legacy-extra/jobs/finance_excel_report.pl.

Aggregates every silver CUSTBILL record of one namespace by currency and record type
(exact DECIMAL, never float), replaces the namespace's rows in ow_tp.gold.finance_billing,
and exports the report to `/Volumes/ow_tp/bronze/landing/<ns>/reports/` twice: a CSV that is
byte-identical to the legacy `finance_billing_<YYYYMMDD>.csv`, and a real `.xlsx` workbook in
place of the legacy CSV-renamed-to-.xls. Distribution is the job-level failure notification;
the run log prints both export paths. Contract:
docs/tech-partnerships/contracts/finance_excel_report.contract.json.

The pure helpers above `main()` have no Spark or dbutils dependency so they can be
imported and unit-tested from scripts/tp_dbx/tests.
"""
from __future__ import annotations

import io
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal

CATALOG = "ow_tp"
SILVER = f"{CATALOG}.silver.custbill_records"
GOLD = f"{CATALOG}.gold.finance_billing"
LANDING = f"/Volumes/{CATALOG}/bronze/landing"
NS_RE = re.compile(r"[a-z0-9][a-z0-9-]{0,31}")
CSV_HEADER = ("Currency", "RecordType", "RecordCount", "TotalAmount")
SHEET_NAME = "finance_billing"
CENT = Decimal("0.01")


def require_ns(ns: str) -> str:
    if not NS_RE.fullmatch(ns or ""):
        raise ValueError(f"ns must match [a-z0-9][a-z0-9-]{{0,31}}: {ns!r}")
    return ns


def resolve_report_date(value: str | None, today: date | None = None) -> date:
    """Widget value `YYYY-MM-DD`; empty/None means the current UTC date (legacy `localtime`)."""
    if value is None or value.strip() == "":
        return today or datetime.now(timezone.utc).date()
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def report_stamp(report_date: date) -> str:
    """Legacy `sprintf("%04d%02d%02d", year, month, day)` filename stamp."""
    return report_date.strftime("%Y%m%d")


def export_paths(ns: str, report_date: date) -> tuple[str, str]:
    stamp = report_stamp(report_date)
    base = f"{LANDING}/{ns}/reports/finance_billing_{stamp}"
    return f"{base}.csv", f"{base}.xlsx"


def record_type_name(rec_type: str) -> str:
    if rec_type == "01":
        return "INVOICE"
    if rec_type == "02":
        return "CREDIT"
    return f"UNKNOWN({rec_type})"


def legacy_sort_key(currency: str, rec_type: str) -> str:
    """The Perl job sorts the string key `"$ccy|$rt"`; `order_rows` must agree with it."""
    return f"{currency}|{rec_type}"


def order_rows(rows: list[tuple]) -> list[tuple]:
    """(currency, rec_type, ...) tuples ordered like `ORDER BY currency, rec_type`."""
    return sorted(rows, key=lambda r: (r[0], r[1]))


def aggregate(records: list[tuple]) -> list[tuple[str, str, int, Decimal]]:
    """Pure reference of the SQL aggregate over (cust_id, bill_amt, currency, rec_type) records:
    empty cust_id rows skipped, exact Decimal sums, ordered by (currency, rec_type)."""
    counts: dict[tuple[str, str], int] = {}
    totals: dict[tuple[str, str], Decimal] = {}
    for cust_id, bill_amt, currency, rec_type in records:
        if cust_id is None or cust_id == "":
            continue
        if bill_amt is None or currency is None or rec_type is None:
            raise ValueError("NULL bill_amt/currency/rec_type reached finance")
        if isinstance(bill_amt, float):
            raise TypeError("bill_amt must be Decimal or str, never float")
        key = (currency, rec_type)
        counts[key] = counts.get(key, 0) + 1
        totals[key] = totals.get(key, Decimal(0)) + Decimal(bill_amt)
    return [
        (currency, rec_type, counts[(currency, rec_type)], totals[(currency, rec_type)].quantize(CENT))
        for currency, rec_type in order_rows(list(counts))
    ]


def format_amount(total: Decimal) -> str:
    """Legacy `%.2f` on the exact total (no float on the way)."""
    if isinstance(total, float):
        raise TypeError("total must be Decimal, never float")
    return f"{Decimal(total).quantize(CENT):f}"


def render_csv(rows: list[tuple]) -> bytes:
    """Header then `%s,%s,%d,%.2f` rows with `\\n` endings, no BOM, one final newline.
    Rows are (currency, record_type_name, record_count, total_amount)."""
    out = [",".join(CSV_HEADER)]
    for currency, record_type, record_count, total in rows:
        out.append(f"{currency},{record_type},{int(record_count):d},{format_amount(total)}")
    return ("\n".join(out) + "\n").encode("utf-8")


def render_xlsx(rows: list[tuple]) -> bytes:
    """One sheet `finance_billing`, same header and rows, amounts numeric with `0.00` format."""
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(list(CSV_HEADER))
    for currency, record_type, record_count, total in rows:
        sheet.append([currency, record_type, int(record_count), None])
        cell = sheet.cell(row=sheet.max_row, column=4)
        cell.value = float(format_amount(total))
        cell.number_format = "0.00"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_rows(aggregated: list[tuple]) -> list[tuple[str, str, int, Decimal]]:
    """(currency, rec_type, count, total) → (currency, record_type_name, count, total)."""
    return [(c, record_type_name(rt), int(n), Decimal(t).quantize(CENT)) for c, rt, n, t in aggregated]


RUN_LOG: list[str] = []


def log(**fields) -> None:
    line = " ".join(f"{key}={value}" for key, value in fields.items())
    RUN_LOG.append(line)
    print(line, flush=True)


def finish(summary: str) -> None:
    dbutils.notebook.exit(json.dumps({"summary": summary, "log": RUN_LOG}))  # noqa: F821


def main() -> None:
    import os

    from pyspark.sql import functions as F
    from pyspark.sql.types import (
        DateType,
        DecimalType,
        LongType,
        StringType,
        StructField,
        StructType,
    )

    ns = require_ns(dbutils.widgets.get("ns"))  # noqa: F821 - injected by Databricks
    report_date = resolve_report_date(dbutils.widgets.get("report_date"))  # noqa: F821
    csv_path, xlsx_path = export_paths(ns, report_date)
    log(stage="finance", ns=ns, report_date=report_date.isoformat(), stamp=report_stamp(report_date))

    nulls = spark.sql(  # noqa: F821
        f"SELECT count(*) FROM {SILVER} WHERE ns = :ns AND cust_id IS NOT NULL AND cust_id <> '' "
        f"AND (bill_amt IS NULL OR currency IS NULL OR rec_type IS NULL)",
        args={"ns": ns},
    ).collect()[0][0]
    if nulls:
        raise RuntimeError(f"{nulls} silver row(s) for ns={ns} carry NULL bill_amt/currency/rec_type")

    aggregated = [
        (r[0], r[1], int(r[2]), Decimal(r[3]))
        for r in spark.sql(  # noqa: F821
            f"SELECT currency, rec_type, count(*) AS record_count, "
            f"CAST(sum(bill_amt) AS DECIMAL(18,2)) AS total_amount "
            f"FROM {SILVER} WHERE ns = :ns AND cust_id IS NOT NULL AND cust_id <> '' "
            f"GROUP BY currency, rec_type ORDER BY currency, rec_type",
            args={"ns": ns},
        ).collect()
    ]
    rows = report_rows(aggregated)
    log(action="aggregated", ns=ns, groups=len(rows), records=sum(r[2] for r in rows))

    spark.sql(f"DELETE FROM {GOLD} WHERE ns = :ns", args={"ns": ns})  # noqa: F821
    if rows:
        schema = StructType([
            StructField("ns", StringType(), False),
            StructField("currency", StringType(), False),
            StructField("record_type", StringType(), False),
            StructField("record_count", LongType(), False),
            StructField("total_amount", DecimalType(18, 2), False),
            StructField("report_date", DateType(), False),
        ])
        (
            spark.createDataFrame(  # noqa: F821
                [(ns, c, rt, n, t, report_date) for c, rt, n, t in rows], schema
            )
            .withColumn("generated_at", F.current_timestamp())
            .write.mode("append")
            .saveAsTable(GOLD)
        )
    log(action="gold_loaded", ns=ns, rows=len(rows))

    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    csv_bytes = render_csv(rows)
    with open(csv_path, "wb") as handle:
        handle.write(csv_bytes)
    log(action="csv_written", path=csv_path, bytes=len(csv_bytes))
    xlsx_bytes = render_xlsx(rows)
    with open(xlsx_path, "wb") as handle:
        handle.write(xlsx_bytes)
    log(action="xlsx_written", path=xlsx_path, bytes=len(xlsx_bytes))

    log(action="done", ns=ns, gold_rows=len(rows), csv=csv_path, xlsx=xlsx_path)
    finish(f"gold_rows={len(rows)} csv={csv_path} xlsx={xlsx_path}")


if __name__ == "__main__":
    main()
